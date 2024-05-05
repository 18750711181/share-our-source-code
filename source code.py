# -*- coding:utf-8 -*-
import os
import time

import openpyxl
from openpyxl.utils import get_column_letter
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
import re
import numpy as np
from scipy.stats import iqr
from openpyxl import Workbook


base_path = ''
original_word = ""
replacement_word = ""
def get_merged_cell_value(sheet, row, col):
    cell = sheet.cell(row, col)
    for range_ in sheet.merged_cells.ranges:
        if cell.coordinate in range_:
            min_row, min_col, max_row, max_col = range_.bounds
            merged_values = []
            for row in sheet.iter_rows(min_row=min_row, min_col=min_col, max_row=max_row, max_col=max_col, values_only=True):
                for value in row:
                    if value is not None:
                        merged_values.append(str(value))
            return ' '.join(merged_values)
    return None

def search_keywords_in_excel_folder(folder_path, keywords, output_file):
    output_workbook = openpyxl.Workbook()
    output_sheet = output_workbook.active
    saved_rows = set()

    for root, _, files in os.walk(folder_path):
        for file_name in files:
            if file_name.endswith('.xlsx'):
                file_path = os.path.join(root, file_name)
                try:
                    wb = openpyxl.load_workbook(file_path, data_only=True)  # Ensure to load values, not formulas
                    sheet = wb.active


                    medicinalproduct_col_index = None
                    for col_num, header_cell in enumerate(sheet[1], start=1):
                        if header_cell.value == 'medicinalproduct':
                            medicinalproduct_col_index = col_num
                            break

                    if medicinalproduct_col_index is not None:
                        for row_num, row in enumerate(sheet.iter_rows(values_only=True), start=1):
                            cell_value = row[medicinalproduct_col_index - 1]
                            if cell_value:
                                cell_data = str(cell_value).lower()
                                merged_value = get_merged_cell_value(sheet, row_num, medicinalproduct_col_index)
                                if merged_value:
                                    cell_data += ' ' + merged_value.lower()

                                if any(keyword.lower() in cell_data for keyword in keywords):
                                    if row not in saved_rows:
                                        row_values = [file_name]
                                        for cell in row:
                                            if isinstance(cell, (str, int, float)):
                                                row_values.append(cell)
                                            else:
                                                row_values.append(str(cell))
                                        output_sheet.append(row_values)
                                        saved_rows.add(row)

                except Exception as e:
                    print("!")


    output_workbook.save(output_file)

folder_path = os.path.join(base_path,"源文件")
keywords_to_search = [f"{original_word}",f"{replacement_word}"]
output_file = os.path.join(base_path,"后续处理",f"{original_word}.xlsx")

search_keywords_in_excel_folder(folder_path, keywords_to_search, output_file)


source_file = os.path.join(base_path, "match.xlsx")
source_df = pd.read_excel(source_file, header=None)


row_to_copy = source_df.iloc[0]


target_file = os.path.join(base_path, "后续处理", f"{original_word}.xlsx")  # 目标文件路径
target_wb = load_workbook(target_file)
target_sheet = target_wb.active


rows_to_insert = list(dataframe_to_rows(pd.DataFrame([row_to_copy]), index=False, header=False))


for r_idx, row in enumerate(rows_to_insert, 1):
    target_sheet.insert_rows(1)
    for c_idx, value in enumerate(row, 1):
        target_sheet.cell(row=1, column=c_idx, value=value)


temp_target_file = os.path.join(base_path, "后续处理", f"temp_{original_word}.xlsx")
target_wb.save(temp_target_file)


target_df = pd.read_excel(temp_target_file)


column_to_replace = 'medicinalproduct'  # 在'medicinalproduct'列进行替换
target_df[column_to_replace] = target_df[column_to_replace].replace(f'{replacement_word}', f'{original_word}', regex=True)

final_target_file = os.path.join(base_path, "后续处理", f"{original_word}-统一关键词和格式更新.xlsx")
target_df.to_excel(final_target_file, index=False)  # 保存到新文件

os.remove(temp_target_file)


file_path = final_target_file
df = pd.read_excel(file_path)

keyword = f'{original_word}'



def extract_data(row):
    drug_charac_list = eval(row['drugcharacterization'])
    med_prod_list = row['medicinalproduct']


    med_prod_names = re.findall(r"'(.*?)'", med_prod_list)


    matching_drugs = [name for name in med_prod_names if keyword.upper() in name.upper()]


    matching_indexes = [med_prod_names.index(drug) for drug in matching_drugs]


    extracted_data = [drug_charac_list[i] for i in matching_indexes]


    if not extracted_data:
        extracted_data = [None]


    return extracted_data[0] if len(extracted_data) > 0 else None



df['drugcharacterization'] = df.apply(extract_data, axis=1)

output_file_path = os.path.join(base_path,"后续处理", f"{original_word}-123.xlsx")
df.to_excel(output_file_path, index=False)




df = pd.read_excel(output_file_path)

df = df[df['drugcharacterization'].isin([2, 3]) == False]

new_directory = os.path.join(base_path, "后续处理", "小包厢")
os.makedirs(new_directory, exist_ok=True)
new_output_file_path = os.path.join(new_directory, f"{original_word}-保留1.xlsx")

df.to_excel(new_output_file_path, index=False)



def 处理Excel文件(file_path, english_pt_dict, matched_pt_set, pt_occurrence_count, matched_pt_list):
    df_dake = pd.read_excel(file_path, engine='openpyxl')  # 指定使用openpyxl引擎读取Excel文件
    print(f"处理文件：{file_path}")

    for index, row in df_dake.iterrows():
        reaction = row['reactionmeddrapt']
        print(f"处理第 {index + 1} 行，不良反应：{reaction}")

        if '[' not in reaction and ']' not in reaction:
            reaction = f"['{reaction}']"

        reaction_cleaned = set(re.findall(r"'(.*?)'", reaction))

        reaction_cleaned = {r.replace('^', "'") for r in reaction_cleaned}

        matched_pt_row = set()

        for r in reaction_cleaned:
            if r in english_pt_dict:
                matched_pt = english_pt_dict[r]
                添加匹配的PT(matched_pt, matched_pt_set, pt_occurrence_count, matched_pt_list, matched_pt_row)
            else:
                print(f"未找到匹配项：{r}")


def 添加匹配的PT(pt, matched_pt_set, pt_occurrence_count, matched_pt_list, matched_pt_row):
    if pt not in matched_pt_set:
        print(f"匹配成功：{pt}")
        matched_pt_set.add(pt)

    if pt not in matched_pt_row:
        matched_pt_row.add(pt)
        if pt in pt_occurrence_count:
            pt_occurrence_count[pt] += 1
        else:
            pt_occurrence_count[pt] = 1
            matched_pt_list.append(pt)


df_pt = pd.read_excel(os.path.join(base_path,'PT匹配.xlsx'))
english_pt_dict = {}
for _, row in df_pt.iterrows():
    english = row['English']
    pt = row['PT']
    if english not in english_pt_dict:
        english_pt_dict[english] = pt


matched_pt_set = set()
pt_occurrence_count = {}
matched_pt_list = []


folder_path = new_directory
for root, _, files in os.walk(folder_path):
    for file in files:
        if file.endswith('.xlsx') and file != 'PT匹配.xlsx':
            file_path = os.path.join(root, file)
            处理Excel文件(file_path, english_pt_dict, matched_pt_set, pt_occurrence_count, matched_pt_list)


result_df = pd.DataFrame({'匹配成功的PT': list(matched_pt_set)})
with pd.ExcelWriter(os.path.join(base_path,"后续处理",f'{original_word}-PT汇总结果.xlsx'), engine='openpyxl') as writer:
    result_df.to_excel(writer, sheet_name='匹配结果', index=False)


count_df = pd.DataFrame(list(pt_occurrence_count.items()), columns=['PT', '出现次数'])
with pd.ExcelWriter(os.path.join(base_path,"后续处理",f'{original_word}-PT汇总结果.xlsx'), engine='openpyxl', mode='a') as writer:
    count_df.to_excel(writer, sheet_name='PT出现次数统计', index=False)








def 处理Excel文件(file_path, english_soc_dict, matched_soc_set, soc_occurrence_count, matched_soc_list):
    df_dake = pd.read_excel(file_path, engine='openpyxl')  # 指定使用openpyxl引擎读取Excel文件


    for index, row in df_dake.iterrows():
        reaction = row['reactionmeddrapt']



        if '[' not in reaction and ']' not in reaction:
            reaction = f"['{reaction}']"


        reaction_cleaned = set(re.findall(r"'(.*?)'", reaction))

        reaction_cleaned = {r.replace('^', "'") for r in reaction_cleaned}

        matched_soc_row = set()

        for r in reaction_cleaned:
            if r in english_soc_dict:
                matched_soc = english_soc_dict[r]
                添加匹配的SOC(matched_soc, matched_soc_set, soc_occurrence_count, matched_soc_list, matched_soc_row)
            else:
                print(1)



def 添加匹配的SOC(soc, matched_soc_set, soc_occurrence_count, matched_soc_list, matched_soc_row):
    if soc not in matched_soc_set:
        print(f"匹配成功：{soc}")
        matched_soc_set.add(soc)

    if soc not in matched_soc_row:
        matched_soc_row.add(soc)
        if soc in soc_occurrence_count:
            soc_occurrence_count[soc] += 1
        else:
            soc_occurrence_count[soc] = 1
            matched_soc_list.append(soc)

df_soc = pd.read_excel(os.path.join(base_path,'1.xlsx'))
english_soc_dict = {}
for _, row in df_soc.iterrows():
    english = row['English']
    soc = row['SOC']
    if english not in english_soc_dict:
        english_soc_dict[english] = soc


matched_soc_set = set()
soc_occurrence_count = {}
matched_soc_list = []


folder_path = new_directory
for root, _, files in os.walk(folder_path):
    for file in files:
        if file.endswith('.xlsx') and file != 'SOC匹配.xlsx':
            file_path = os.path.join(root, file)
            处理Excel文件(file_path, english_soc_dict, matched_soc_set, soc_occurrence_count, matched_soc_list)


result_df = pd.DataFrame({'匹配成功的SOC': list(matched_soc_set)})
with pd.ExcelWriter(os.path.join(base_path,"后续处理", f'{original_word}-SOC汇总结果.xlsx'), engine='openpyxl') as writer:
    result_df.to_excel(writer, sheet_name='匹配结果', index=False)


count_df = pd.DataFrame(list(soc_occurrence_count.items()), columns=['SOC', '出现次数'])
with pd.ExcelWriter(os.path.join(base_path,"后续处理", f'{original_word}-SOC汇总结果.xlsx'), engine='openpyxl', mode='a') as writer:
    count_df.to_excel(writer, sheet_name='SOC出现次数统计', index=False)





def 处理Excel文件(file_path, english_pt_dict, matched_pt_set, pt_occurrence_count, matched_pt_list):
    df_dake = pd.read_excel(file_path, engine='openpyxl')  # 指定使用openpyxl引擎读取Excel文件
    print(f"处理文件：{file_path}")

    for index, row in df_dake.iterrows():
        reaction = row['reactionmeddrapt']



        if '[' not in reaction and ']' not in reaction:
            reaction = f"['{reaction}']"


        reaction_cleaned = set(re.findall(r"'(..*?)'", reaction))


        reaction_cleaned = {r.replace('^', "'") for r in reaction_cleaned}

        matched_pt_row = set()

        for r in reaction_cleaned:
            if r in english_pt_dict:
                matched_pt = english_pt_dict[r]
                添加匹配的PT(matched_pt, matched_pt_set, pt_occurrence_count, matched_pt_list, matched_pt_row)
            else:
                print(f"未找到匹配项：{r}")


def 添加匹配的PT(pt, matched_pt_set, pt_occurrence_count, matched_pt_list, matched_pt_row):
    if pt not in matched_pt_set:
        print(f"匹配成功：{pt}")
        matched_pt_set.add(pt)

    if pt not in matched_pt_row:
        matched_pt_row.add(pt)
        if pt in pt_occurrence_count:
            pt_occurrence_count[pt] += 1
        else:
            pt_occurrence_count[pt] = 1
            matched_pt_list.append(pt)


df_pt = pd.read_excel(os.path.join(base_path,'PT匹配.xlsx'))
english_pt_dict = {}
for _, row in df_pt.iterrows():
    english = row['English']
    pt = row['PT']
    if english not in english_pt_dict:
        english_pt_dict[english] = pt


matched_pt_set = set()
pt_occurrence_count = {}
matched_pt_list = []


folder_path = os.path.join(base_path,"源文件")
for root, _, files in os.walk(folder_path):
    for file in files:
        if file.endswith('.xlsx') and file != 'PT匹配.xlsx':
            file_path = os.path.join(root, file)
            处理Excel文件(file_path, english_pt_dict, matched_pt_set, pt_occurrence_count, matched_pt_list)


result_df = pd.DataFrame({'匹配成功的PT': list(matched_pt_set)})
with pd.ExcelWriter(os.path.join(base_path,'PT汇总结果.xlsx'), engine='openpyxl') as writer:
    result_df.to_excel(writer, sheet_name='匹配结果', index=False)

count_df = pd.DataFrame(list(pt_occurrence_count.items()), columns=['PT', '出现次数'])
with pd.ExcelWriter(os.path.join(base_path,'PT汇总结果.xlsx'), engine='openpyxl', mode='a') as writer:
    count_df.to_excel(writer, sheet_name='PT出现次数统计', index=False)

print(f"\n共匹配成功了 {len(matched_pt_set)} 条数据。")
print("结果已保存在汇总结果.xlsx中。")



def 处理Excel文件(file_path, english_soc_dict, matched_soc_set, soc_occurrence_count, matched_soc_list):
    df_dake = pd.read_excel(file_path, engine='openpyxl')
    print(f"处理文件：{file_path}")

    for index, row in df_dake.iterrows():
        reaction = row['reactionmeddrapt']
        print(f"处理第 {index + 1} 行，不良反应：{reaction}")


        if '[' not in reaction and ']' not in reaction:
            reaction = f"['{reaction}']"


        reaction_cleaned = set(re.findall(r"'(..*?)'", reaction))

        reaction_cleaned = {r.replace('^^', "'") for r in reaction_cleaned}

        matched_soc_row = set()

        for r in reaction_cleaned:
            if r in english_soc_dict:
                matched_soc = english_soc_dict[r]
                添加匹配的SOC(matched_soc, matched_soc_set, soc_occurrence_count, matched_soc_list, matched_soc_row)
            else:
                print(f"未找到匹配项：{r}")


def 添加匹配的SOC(soc, matched_soc_set, soc_occurrence_count, matched_soc_list, matched_soc_row):
    if soc not in matched_soc_set:
        print(f"匹配成功：{soc}")
        matched_soc_set.add(soc)

    if soc not in matched_soc_row:
        matched_soc_row.add(soc)
        if soc in soc_occurrence_count:
            soc_occurrence_count[soc] += 1
        else:
            soc_occurrence_count[soc] = 1
            matched_soc_list.append(soc)


df_soc = pd.read_excel(os.path.join(base_path,'SOC匹配.xlsx'))
english_soc_dict = {}
for _, row in df_soc.iterrows():
    english = row['English']
    soc = row['SOC']
    if english not in english_soc_dict:
        english_soc_dict[english] = soc


matched_soc_set = set()
soc_occurrence_count = {}
matched_soc_list = []


folder_path = os.path.join(base_path,"源文件")
for root, _, files in os.walk(folder_path):
    for file in files:
        if file.endswith('.xlsx') and file != 'SOC匹配.xlsx':
            file_path = os.path.join(root, file)
            处理Excel文件(file_path, english_soc_dict, matched_soc_set, soc_occurrence_count, matched_soc_list)


result_df = pd.DataFrame({'匹配成功的SOC': list(matched_soc_set)})
with pd.ExcelWriter(os.path.join(base_path, f'{original_word}-SOC汇总结果.xlsx'), engine='openpyxl') as writer:
    result_df.to_excel(writer, sheet_name='匹配结果', index=False)


count_df = pd.DataFrame(list(soc_occurrence_count.items()), columns=['SOC', '出现次数'])
with pd.ExcelWriter(os.path.join(base_path, f'{original_word}-SOC汇总结果.xlsx'), engine='openpyxl', mode='a') as writer:
    count_df.to_excel(writer, sheet_name='SOC出现次数统计', index=False)








input_file_path = new_output_file_path

output_file_path = os.path.join(new_directory,f'{original_word}-部分临床特征统计.xlsx')


df = pd.read_excel(input_file_path)


columns_to_count = [
    'reportercountry', 'patientsex', 'qualification', 'serious',
    'seriousnesscongenitalanomali', 'seriousnessdeath',
    'seriousnessdisabling', 'seriousnesshospitalization',
    'seriousnesslifethreatening', 'seriousnessother',
    'YEAR', 'patientweight'
]


wb = Workbook()


for column in columns_to_count:

    if column not in df:
        continue


    counts = df[column].value_counts().reset_index()
    counts.columns = [column, 'Frequency']


    total_frequency = counts['Frequency'].sum()

    total_row = pd.DataFrame({column: ['Total'], 'Frequency': [total_frequency]})
    counts = pd.concat([counts, total_row], ignore_index=True)


    ws = wb.create_sheet(title=column[:31])


    for r in dataframe_to_rows(counts, index=False, header=True):
        ws.append(r)


    if column == 'patientweight':

        valid_weights = df[column].dropna()
        median_weight = np.median(valid_weights)
        iqr_weight = iqr(valid_weights)


        summary_df = pd.DataFrame({
            'Statistic': ['Median', 'IQR (Q1-Q3)'],
            'Value': [median_weight, iqr_weight]
        })

        ws_summary = wb.create_sheet(title=column[:31] + '_Summary')


        for r in dataframe_to_rows(summary_df, index=False, header=True):
            ws_summary.append(r)

        weight_categories = pd.cut(valid_weights, bins=[0, 80, 100, np.inf], right=False, labels=['<80', '80-100', '>100'])
        category_counts = weight_categories.value_counts().reset_index()
        category_counts.columns = ['Weight Category', 'Frequency']

        ws_weight_cat = wb.create_sheet(title=column[:31] + '_Weight_Categories')
        for r in dataframe_to_rows(category_counts, index=False, header=True):
            ws_weight_cat.append(r)

if 'Sheet' in wb.sheetnames:
    del wb['Sheet']


wb.save(output_file_path)

print(f"Counted data has been saved to '{output_file_path}'.")

